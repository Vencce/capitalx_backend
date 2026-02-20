from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from datetime import datetime

from .models import Administradora, Carta, Configuracao
from .serializers import AdministradoraSerializer, CartaSerializer, ConfiguracaoSerializer
from .services import executar_sincronizacao

class AdministradoraViewSet(viewsets.ModelViewSet):
    queryset = Administradora.objects.all()
    serializer_class = AdministradoraSerializer
    permission_classes = [AllowAny] 

class CartaViewSet(viewsets.ModelViewSet):
    queryset = Carta.objects.all()
    serializer_class = CartaSerializer
    
    def get_queryset(self):
        queryset = Carta.objects.all()
        # Permite filtrar por origem via URL: /api/cartas/?origem=PARCEIRO
        origem = self.request.query_params.get('origem')
        if origem:
            queryset = queryset.filter(origem=origem)
        return queryset

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [AllowAny()]
        return [IsAuthenticated()]

class SincronizarCartasView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        resultado = executar_sincronizacao()
        if resultado:
            return Response(resultado)
        return Response({"error": "Falha na sincronização"}, status=400)

class ConfiguracaoView(APIView):
    def get_permissions(self):
        if self.request.method == 'GET': return [AllowAny()]
        return [IsAuthenticated()]

    def get(self, request):
        config, _ = Configuracao.objects.get_or_create(id=1)
        return Response(ConfiguracaoSerializer(config).data)

    def post(self, request):
        config, _ = Configuracao.objects.get_or_create(id=1)
        serializer = ConfiguracaoSerializer(config, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

class CustomLoginView(ObtainAuthToken):
    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        token = Token.objects.get(key=response.data['token'])
        return Response({'token': token.key, 'user_id': token.user_id})

class ExportarExcelView(APIView):
    def get(self, request):
        # 1. Filtra os dados (Mesma lógica da busca)
        queryset = Carta.objects.all()
        
        tipo = request.query_params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo)
            
        # 2. Configura o Arquivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"cartas_capitalx_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Cartas Contempladas"

        # 3. Cabeçalho Estilizado
        headers = [
            'Cód. Cota', 'Categoria', 'Crédito', 'Entrada', 
            'Nº Parcelas', 'Vlr Parcela', 'Saldo Devedor', 
            'Taxa Transf.', 'Administradora', 'Status'
        ]
        ws.append(headers)

        # Estilo do cabeçalho (Azul escuro com texto branco)
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # 4. Dados
        for carta in queryset:
            ws.append([
                carta.codigo,
                carta.get_tipo_display(),
                f"R$ {carta.valor_credito:,.2f}",
                f"R$ {carta.valor_entrada:,.2f}",
                carta.numero_parcelas,
                f"R$ {carta.valor_parcela:,.2f}",
                f"R$ {carta.saldo_devedor:,.2f}" if carta.saldo_devedor else "-",
                str(carta.taxa_transferencia),
                carta.administradora.nome if carta.administradora else '-',
                carta.get_status_display()
            ])

        # Ajuste automático de largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(response)
        return response

class ExportarPDFView(APIView):
    def get(self, request):
        # 1. Filtros
        queryset = Carta.objects.all()
        tipo = request.query_params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo)

        # 2. Configura PDF (Paisagem para caber mais colunas)
        response = HttpResponse(content_type='application/pdf')
        filename = f"cartas_capitalx_{datetime.now().strftime('%d-%m-%Y')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        p = canvas.Canvas(response, pagesize=landscape(A4))
        width, height = landscape(A4)
        y = height - 50

        # Cabeçalho do Documento
        p.setFont("Helvetica-Bold", 18)
        p.setFillColorRGB(0.12, 0.23, 0.54) # Azul CapitalX
        p.drawString(30, y, "Relatório de Cartas Contempladas")
        
        p.setFont("Helvetica", 10)
        p.setFillColorRGB(0, 0, 0)
        p.drawString(30, y - 20, f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
        y -= 50

        # Cabeçalho da Tabela
        headers = [
            ("CÓD.", 30), ("CATEGORIA", 80), ("ADM", 160), ("CRÉDITO", 250), 
            ("ENTRADA", 330), ("PARCELAS", 410), ("SALDO DEV.", 500), ("STATUS", 600)
        ]
        
        p.setFont("Helvetica-Bold", 9)
        for title, x in headers:
            p.drawString(x, y, title)
        
        y -= 10
        p.line(30, y, 800, y) # Linha separadora
        y -= 20

        # Dados
        p.setFont("Helvetica", 9)
        for carta in queryset:
            if y < 50: # Nova página
                p.showPage()
                y = height - 50
                p.setFont("Helvetica-Bold", 9) # Repete cabeçalho
                for title, x in headers:
                    p.drawString(x, y, title)
                y -= 20
                p.setFont("Helvetica", 9)
            
            # Formatação de valores
            credito = f"R$ {carta.valor_credito:,.2f}"
            entrada = f"R$ {carta.valor_entrada:,.2f}"
            parcelas = f"{carta.numero_parcelas}x R$ {carta.valor_parcela:,.2f}"
            saldo = f"R$ {carta.saldo_devedor:,.2f}" if carta.saldo_devedor else "-"

            p.drawString(30, y, str(carta.codigo))
            p.drawString(80, y, carta.get_tipo_display())
            p.drawString(160, y, carta.administradora.nome[:15]) # Trunca nome longo
            p.drawString(250, y, credito)
            p.drawString(330, y, entrada)
            p.drawString(410, y, parcelas)
            p.drawString(500, y, saldo)
            
            # Status colorido (simulado via texto)
            p.drawString(600, y, carta.status)
            
            y -= 15 # Espaço entre linhas

        p.showPage()
        p.save()
        return response